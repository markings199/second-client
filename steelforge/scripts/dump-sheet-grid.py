"""
Dump small/medium sheets as a markdown table (grid layout) so the layout is
readable like the original Excel sheet.

Each cell is shown as either:
  • a static label / number
  • a formula prefixed with '=' (and we also emit a separate 'cached value' note below the table)

Run:
  python steelforge/scripts/dump-sheet-grid.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "reference" / "workbook.xlsx"

# Sheets to dump as full grids (skip huge catalog sheets).
SHEETS = [
    "TENSION",
    "TENSION RODS",
    "BENDING",
    "SHEAR DESIGN",
    "SHEAR ANALYSIS",
    "TENSION SHAPES",
    "S(ASTM)",
    "SECTION ZX",
]
# Sheets to dump only their "compute" area (top N rows), skipping the catalog tail.
SHEETS_TOP = {
    "COMPRESSION": 80,
}
OUT = ROOT / "reference" / "workbook-grids.md"


def fmt(v):
    if v is None:
        return ""
    # openpyxl returns ArrayFormula objects for {=...} formulas — pull out the text.
    if hasattr(v, "text"):
        v = v.text
    if isinstance(v, float):
        return f"{v:.6g}"
    return str(v).replace("|", "\\|").replace("\n", " ")


def main():
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    wb_v = openpyxl.load_workbook(XLSX, data_only=True)

    lines = ["# Workbook small-sheet grid dump", ""]
    lines.append("Each cell shows the *raw content* (formula if any, else label/number).")
    lines.append("A note table below each grid shows the cached value for any formula cell.")
    lines.append("")

    sheet_jobs = [(s, None) for s in SHEETS] + [(s, cap) for s, cap in SHEETS_TOP.items()]
    for name, top_row_cap in sheet_jobs:
        if name not in wb_f.sheetnames:
            continue
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        max_r = ws_f.max_row or 0
        max_c = ws_f.max_column or 0
        if top_row_cap is not None:
            max_r = min(max_r, top_row_cap)
        # Clip empty trailing rows/cols to keep things readable.
        # Find last non-empty row and col actually used.
        last_r = 0
        last_c = 0
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                val = ws_f.cell(row=r, column=c).value
                if val is not None:
                    if r > last_r:
                        last_r = r
                    if c > last_c:
                        last_c = c
        # cap to keep doc reasonable
        last_r = min(last_r, top_row_cap if top_row_cap is not None else 70)
        last_c = min(last_c, 30)

        lines.append(f"## Sheet: `{name}`  ({last_r} rows × {last_c} cols)")
        lines.append("")
        # Header row of column letters.
        hdr = ["row"] + [get_column_letter(c) for c in range(1, last_c + 1)]
        sep = ["---"] * len(hdr)
        lines.append("| " + " | ".join(hdr) + " |")
        lines.append("| " + " | ".join(sep) + " |")

        cached = []
        for r in range(1, last_r + 1):
            row_cells = [str(r)]
            for c in range(1, last_c + 1):
                cf = ws_f.cell(row=r, column=c)
                v_raw = cf.value
                if v_raw is None:
                    row_cells.append("")
                    continue
                if hasattr(v_raw, "text"):
                    v_raw = v_raw.text
                if isinstance(v_raw, str) and v_raw.startswith("="):
                    cv = ws_v.cell(row=r, column=c).value
                    row_cells.append("`" + fmt(v_raw) + "`")
                    if cv is not None:
                        cached.append((cf.coordinate, fmt(cv)))
                else:
                    row_cells.append(fmt(v_raw))
            lines.append("| " + " | ".join(row_cells) + " |")
        lines.append("")
        if cached:
            lines.append("### Cached values for formula cells")
            lines.append("")
            lines.append("| Cell | Cached value |")
            lines.append("| --- | --- |")
            for addr, val in cached:
                lines.append(f"| {addr} | {val} |")
            lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()

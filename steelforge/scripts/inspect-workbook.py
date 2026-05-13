"""
Dump every sheet of the client workbook into a Markdown report so the
implementing agent can re-create each calculation in JavaScript.

For every cell:
  - skips empty cells
  - shows address, formula (if any), and evaluated value (with data_only=True)

Run:
  python steelforge/scripts/inspect-workbook.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "reference" / "workbook.xlsx"
OUT_MD = ROOT / "reference" / "workbook-formulas.md"
OUT_JSON = ROOT / "reference" / "workbook-formulas.json"


def stringify(v):
    if v is None:
        return ""
    if isinstance(v, float):
        # keep precision but trim noise
        s = f"{v:.10g}"
        return s
    return str(v)


def is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def main():
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)  # formulas
    wb_v = openpyxl.load_workbook(XLSX, data_only=True)   # cached values

    summary = []
    md_lines = ["# Client workbook formula dump", ""]
    md_lines.append(f"Source: `{XLSX.relative_to(ROOT)}`")
    md_lines.append("")
    md_lines.append(f"Sheets ({len(wb_f.sheetnames)}): {', '.join(wb_f.sheetnames)}")
    md_lines.append("")

    all_data = {}

    for name in wb_f.sheetnames:
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        rows_used = 0
        cols_used = 0
        cells = []

        max_row = ws_f.max_row or 0
        max_col = ws_f.max_column or 0

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cf = ws_f.cell(row=r, column=c)
                cv = ws_v.cell(row=r, column=c)
                fval = cf.value
                vval = cv.value
                if fval is None and vval is None:
                    continue
                addr = cf.coordinate
                is_f = is_formula(fval)
                cells.append(
                    {
                        "addr": addr,
                        "row": r,
                        "col": c,
                        "formula": fval if is_f else None,
                        "value": stringify(vval if vval is not None else fval),
                        "raw": None if is_f else stringify(fval),
                    }
                )
                rows_used = max(rows_used, r)
                cols_used = max(cols_used, c)

        all_data[name] = {
            "rows": rows_used,
            "cols": cols_used,
            "cells": cells,
        }

        formula_cells = [c for c in cells if c["formula"]]

        md_lines.append(f"## Sheet: `{name}`  ({rows_used} rows × {cols_used} cols, {len(formula_cells)} formula cells)")
        md_lines.append("")

        # Compact "formulas" table (just formula cells with cached value)
        if formula_cells:
            md_lines.append("### Formulas")
            md_lines.append("")
            md_lines.append("| Cell | Formula | Cached value |")
            md_lines.append("| --- | --- | --- |")
            for c in formula_cells:
                f = c["formula"].replace("|", "\\|")
                v = c["value"].replace("|", "\\|")
                md_lines.append(f"| {c['addr']} | `{f}` | {v} |")
            md_lines.append("")

        # Compact "labels & constants" table (top 60 non-formula text/number cells, for layout context)
        text_cells = [c for c in cells if not c["formula"] and c["value"] != ""]
        if text_cells:
            md_lines.append("### Labels & static values (first 80 non-formula cells)")
            md_lines.append("")
            md_lines.append("| Cell | Value |")
            md_lines.append("| --- | --- |")
            for c in text_cells[:80]:
                v = c["value"].replace("|", "\\|")
                md_lines.append(f"| {c['addr']} | {v} |")
            if len(text_cells) > 80:
                md_lines.append(f"| … | ({len(text_cells) - 80} more) |")
            md_lines.append("")

        summary.append((name, rows_used, cols_used, len(formula_cells), len(text_cells)))

    md_lines.insert(4, "")
    md_lines.insert(4, "")
    md_lines.insert(4, "| Sheet | Rows | Cols | Formulas | Text/const cells |")
    md_lines.insert(5, "| --- | ---: | ---: | ---: | ---: |")
    for name, rr, cc, ff, tt in summary:
        md_lines.insert(6, f"| `{name}` | {rr} | {cc} | {ff} | {tt} |")

    OUT_MD.write_text("\n".join(md_lines), encoding="utf-8")
    OUT_JSON.write_text(json.dumps(all_data, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_MD}")
    print(f"Wrote {OUT_JSON}")
    for name, rr, cc, ff, tt in summary:
        print(f"  {name}: {ff} formulas, {tt} text/const, {rr}r×{cc}c")


if __name__ == "__main__":
    main()
